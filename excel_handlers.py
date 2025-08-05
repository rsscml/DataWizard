import pandas as pd
import numpy as np
from openpyxl import load_workbook
import re
from typing import Tuple, Dict, List, Optional, Any
import logging

# Set up logging
logger = logging.getLogger(__name__)

def find_data_start_position(file_path: str, sheet_name: str, max_search_rows: int = 20, max_search_cols: int = 10) -> Tuple[int, int, bool]:
    """
    Find where the actual tabular data starts in a worksheet.
    
    Returns:
        (start_row, start_col, found): Row and column indices (0-based) where data starts, and success flag
    """
    try:
        # Use openpyxl to examine the raw structure
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook[sheet_name]
        
        best_position = (0, 0)
        best_score = 0
        
        # Search in a grid pattern for the best data start position
        for start_row in range(min(max_search_rows, worksheet.max_row)):
            for start_col in range(min(max_search_cols, worksheet.max_column)):
                score = evaluate_data_start_position(worksheet, start_row, start_col)
                if score > best_score:
                    best_score = score
                    best_position = (start_row, start_col)
        
        workbook.close()
        
        # If we found a position with a reasonable score, use it
        if best_score > 0.3:  # Threshold for "good enough" data start
            return best_position[0], best_position[1], True
        else:
            return 0, 0, False
            
    except Exception as e:
        logger.warning(f"Error finding data start position for {sheet_name}: {e}")
        return 0, 0, False

def evaluate_data_start_position(worksheet, start_row: int, start_col: int) -> float:
    """
    Evaluate how good a position is as a data start point.
    Returns a score between 0 and 1.
    """
    try:
        # Look at a sample region (10x10) starting from this position
        sample_rows = 10
        sample_cols = 10
        
        total_cells = 0
        filled_cells = 0
        header_like_cells = 0
        data_like_cells = 0
        
        # Check the first row (potential headers)
        first_row_values = []
        for col in range(start_col, min(start_col + sample_cols, worksheet.max_column)):
            cell = worksheet.cell(row=start_row + 1, column=col + 1)  # openpyxl is 1-indexed
            value = cell.value
            first_row_values.append(value)
            
            if value is not None:
                filled_cells += 1
                if isinstance(value, str) and len(value.strip()) > 0:
                    header_like_cells += 1
            total_cells += 1
        
        # Check subsequent rows (potential data)
        for row in range(start_row + 1, min(start_row + sample_rows, worksheet.max_row)):
            for col in range(start_col, min(start_col + sample_cols, worksheet.max_column)):
                cell = worksheet.cell(row=row + 1, column=col + 1)
                value = cell.value
                
                if value is not None:
                    filled_cells += 1
                    if isinstance(value, (int, float)) or (isinstance(value, str) and len(value.strip()) > 0):
                        data_like_cells += 1
                total_cells += 1
        
        if total_cells == 0:
            return 0
        
        # Calculate score based on multiple factors
        fill_ratio = filled_cells / total_cells
        header_ratio = header_like_cells / len(first_row_values) if first_row_values else 0
        data_ratio = data_like_cells / max(1, total_cells - len(first_row_values))
        
        # Bonus for having reasonable headers
        header_quality = 0
        if first_row_values:
            unique_headers = len(set(str(v).strip() for v in first_row_values if v is not None))
            total_headers = len([v for v in first_row_values if v is not None])
            if total_headers > 0:
                header_quality = unique_headers / total_headers
        
        # Combined score
        score = (fill_ratio * 0.3 + header_ratio * 0.3 + data_ratio * 0.3 + header_quality * 0.1)
        
        return min(1.0, score)
        
    except Exception as e:
        logger.warning(f"Error evaluating position ({start_row}, {start_col}): {e}")
        return 0

def detect_and_handle_merged_headers(file_path: str, sheet_name: str, header_row: int = 0) -> List[str]:
    """
    Detect merged cells in headers and create appropriate column names.
    
    Returns:
        List of cleaned column names
    """
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook[sheet_name]
        
        # Get merged cell ranges
        merged_ranges = worksheet.merged_cells.ranges
        
        # Create a mapping of cell positions to their merged values
        merged_cell_map = {}
        for merged_range in merged_ranges:
            # Get the value from the top-left cell of the merged range
            top_left_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
            value = top_left_cell.value
            
            # Apply this value to all cells in the merged range
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cell_map[(row, col)] = value
        
        # Extract headers from the specified row
        headers = []
        max_col = worksheet.max_column
        
        for col in range(1, max_col + 1):
            cell_pos = (header_row + 1, col)  # openpyxl is 1-indexed
            
            if cell_pos in merged_cell_map:
                value = merged_cell_map[cell_pos]
            else:
                cell = worksheet.cell(row=header_row + 1, column=col)
                value = cell.value
            
            # Clean up the header value
            if value is None:
                headers.append(f"Column_{col}")
            else:
                clean_header = str(value).strip()
                if not clean_header:
                    clean_header = f"Column_{col}"
                headers.append(clean_header)
        
        workbook.close()
        return headers
        
    except Exception as e:
        logger.warning(f"Error handling merged headers for {sheet_name}: {e}")
        # Fallback: return generic column names
        try:
            df_temp = pd.read_excel(file_path, sheet_name=sheet_name, nrows=1)
            return [f"Column_{i+1}" for i in range(len(df_temp.columns))]
        except:
            return []

def clean_and_flatten_headers(headers: List[Any], multi_level_separator: str = "_") -> List[str]:
    """
    Clean and flatten multi-level headers.
    
    Args:
        headers: List of header values (could be tuples for multi-level)
        multi_level_separator: Separator to use when flattening multi-level headers
    
    Returns:
        List of cleaned, flat header strings
    """
    cleaned_headers = []
    
    for i, header in enumerate(headers):
        if isinstance(header, tuple):
            # Multi-level header - flatten it
            parts = []
            for part in header:
                if part is not None and str(part).strip() and not str(part).startswith('Unnamed'):
                    parts.append(str(part).strip())
            
            if parts:
                clean_header = multi_level_separator.join(parts)
            else:
                clean_header = f"Column_{i+1}"
        else:
            # Single-level header
            if header is None or str(header).strip() == '' or str(header).startswith('Unnamed'):
                clean_header = f"Column_{i+1}"
            else:
                clean_header = str(header).strip()
        
        # Ensure uniqueness
        original_header = clean_header
        counter = 1
        while clean_header in cleaned_headers:
            clean_header = f"{original_header}_{counter}"
            counter += 1
        
        cleaned_headers.append(clean_header)
    
    return cleaned_headers

def extract_data_region(file_path: str, sheet_name: str, start_row: int = 0, start_col: int = 0) -> pd.DataFrame:
    """
    Extract the data region from a worksheet starting from specified position.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Name of the worksheet
        start_row: Row index to start reading from (0-based)
        start_col: Column index to start reading from (0-based)
    
    Returns:
        Cleaned DataFrame
    """
    try:
        # Read the data starting from the specified position
        df = pd.read_excel(
            file_path, 
            sheet_name=sheet_name, 
            skiprows=start_row,
            usecols=lambda col: col >= start_col if isinstance(col, int) else True
        )
        
        # Handle potential merged/multi-level headers
        if df.empty:
            return df
        
        # Clean column names
        if hasattr(df.columns, 'nlevels') and df.columns.nlevels > 1:
            # Multi-level columns
            df.columns = clean_and_flatten_headers(df.columns.tolist())
        else:
            # Single-level columns
            df.columns = clean_and_flatten_headers(df.columns.tolist())
        
        # Remove completely empty rows and columns
        df = df.dropna(how='all').dropna(axis=1, how='all')
        
        # Reset index
        df = df.reset_index(drop=True)
        
        return df
        
    except Exception as e:
        logger.warning(f"Error extracting data region from {sheet_name}: {e}")
        # Fallback to standard reading
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            df.columns = clean_and_flatten_headers(df.columns.tolist())
            return df.dropna(how='all').dropna(axis=1, how='all').reset_index(drop=True)
        except:
            return pd.DataFrame()

def attempt_worksheet_recovery(file_path: str, sheet_name: str) -> Tuple[bool, pd.DataFrame, List[str]]:
    """
    Attempt to recover a worksheet that failed initial validation.
    
    Returns:
        (success, dataframe, recovery_notes): Success flag, recovered DataFrame, and notes about what was done
    """
    recovery_notes = []
    
    try:
        # Step 1: Find the optimal data start position
        start_row, start_col, found_start = find_data_start_position(file_path, sheet_name)
        
        if not found_start:
            recovery_notes.append("Could not locate data start position")
            return False, pd.DataFrame(), recovery_notes
        
        if start_row > 0 or start_col > 0:
            recovery_notes.append(f"Data found starting at row {start_row+1}, column {start_col+1} (not A1)")
        
        # Step 2: Extract data from the identified region
        df = extract_data_region(file_path, sheet_name, start_row, start_col)
        
        if df.empty:
            recovery_notes.append("No data found in identified region")
            return False, pd.DataFrame(), recovery_notes
        
        # Step 3: Handle merged headers if detected
        try:
            cleaned_headers = detect_and_handle_merged_headers(file_path, sheet_name, start_row)
            if cleaned_headers and len(cleaned_headers) == len(df.columns):
                df.columns = cleaned_headers
                recovery_notes.append("Merged header cells detected and handled")
        except Exception as e:
            recovery_notes.append(f"Header processing had issues: {e}")
        
        # Step 4: Final validation of recovered data
        if df.shape[0] < 2:
            recovery_notes.append("Insufficient rows after recovery")
            return False, pd.DataFrame(), recovery_notes
        
        if df.shape[1] < 1:
            recovery_notes.append("No columns after recovery")
            return False, pd.DataFrame(), recovery_notes
        
        # Check data density
        total_cells = df.shape[0] * df.shape[1]
        non_null_cells = df.count().sum()
        data_density = non_null_cells / total_cells if total_cells > 0 else 0
        
        if data_density < 0.05:  # Less than 5% data
            recovery_notes.append(f"Data too sparse after recovery: {data_density:.1%} filled")
            return False, pd.DataFrame(), recovery_notes
        
        recovery_notes.append(f"Successfully recovered {df.shape[0]}×{df.shape[1]} data with {data_density:.1%} density")
        
        # Clean data types for JSON serialization (same as original code)
        for col in df.columns:
            dtype_name = str(df[col].dtype)
            if any(problematic in dtype_name for problematic in
                   ['Int64', 'Float64', 'boolean', 'string']):
                recovery_notes.append(f"Converting column '{col}' from {dtype_name} to standard type")
                if 'Int' in dtype_name:
                    df[col] = pd.to_numeric(df[col], errors='coerce').astype('float64')
                elif 'Float' in dtype_name:
                    df[col] = pd.to_numeric(df[col], errors='coerce').astype('float64')
                elif 'boolean' in dtype_name:
                    df[col] = df[col].astype('object')
                else:  # string type
                    df[col] = df[col].astype('object')
        
        return True, df, recovery_notes
        
    except Exception as e:
        recovery_notes.append(f"Recovery failed with error: {e}")
        return False, pd.DataFrame(), recovery_notes

def is_worksheet_properly_formatted(file_path, sheet_name):
    """
    Enhanced version: Check if a worksheet is properly formatted, with recovery attempts.
    Returns: (is_valid, data_frame, issues_found, recovery_info)
    """
    issues = []
    recovery_info = {}
    
    try:
        # Try standard pandas reading first
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        # Check 1: Must not be empty
        if df.empty:
            issues.append("Worksheet is empty")
            
            # Attempt recovery
            recovery_success, recovered_df, recovery_notes = attempt_worksheet_recovery(file_path, sheet_name)
            if recovery_success:
                recovery_info = {
                    'attempted': True,
                    'successful': True,
                    'notes': recovery_notes,
                    'original_issues': issues.copy()
                }
                return True, recovered_df, [], recovery_info
            else:
                recovery_info = {
                    'attempted': True,
                    'successful': False,
                    'notes': recovery_notes
                }
                return False, None, issues, recovery_info
        
        # Continue with original checks, but be less strict
        original_df = df.copy()
        
        # Check 2: Must have reasonable dimensions
        if df.shape[0] < 2:  # Need at least 2 rows (header + data)
            issues.append(f"Too few rows: {df.shape[0]} (need at least 2)")
        
        if df.shape[1] < 1:  # Need at least 1 column
            issues.append(f"No columns found")
        
        # Check 3: Column names should be reasonable (less strict than before)
        unnamed_cols = [col for col in df.columns if str(col).startswith('Unnamed')]
        unnamed_ratio = len(unnamed_cols) / len(df.columns) if len(df.columns) > 0 else 1
        
        if unnamed_ratio > 0.8:  # More lenient: 80% instead of 50%
            issues.append(f"Too many unnamed columns: {len(unnamed_cols)}/{len(df.columns)}")
        
        # Check 4: Should have some actual data (not all NaN)
        non_null_data = df.count().sum()  # Total non-null values across all columns
        total_cells = df.shape[0] * df.shape[1]
        
        if non_null_data == 0:
            issues.append("No actual data found (all cells empty)")
        elif non_null_data / total_cells < 0.05:  # More lenient: 5% instead of 10%
            issues.append(f"Very sparse data: only {non_null_data}/{total_cells} cells have data")
        
        # Check 5: First column analysis (less strict)
        if df.shape[1] > 0:
            first_col_filled = df.iloc[:, 0].count()
            if first_col_filled / len(df) < 0.1:  # More lenient: 10% instead of 30%
                issues.append("First column mostly empty - data may not start at A1")
        
        # Check 6: Column names should be reasonably unique (less strict)
        if df.shape[1] > 0:
            unique_cols = len(set(str(col) for col in df.columns))
            if unique_cols / len(df.columns) < 0.6:  # More lenient: 60% instead of 80%
                issues.append("Many duplicate column names")
        
        # If we have issues, attempt recovery
        if issues:
            print(f"  Worksheet '{sheet_name}' has issues: {issues}")
            print(f"  Attempting recovery...")
            
            recovery_success, recovered_df, recovery_notes = attempt_worksheet_recovery(file_path, sheet_name)
            
            if recovery_success:
                recovery_info = {
                    'attempted': True,
                    'successful': True,
                    'notes': recovery_notes,
                    'original_issues': issues.copy()
                }
                print(f"  Recovery successful: {recovery_notes}")
                return True, recovered_df, [], recovery_info
            else:
                recovery_info = {
                    'attempted': True,
                    'successful': False,
                    'notes': recovery_notes
                }
                print(f"  Recovery failed: {recovery_notes}")
                return False, None, issues, recovery_info
        
        # All checks passed without issues
        print(f"  Worksheet '{sheet_name}' is properly formatted: {df.shape[0]}×{df.shape[1]}")
        recovery_info = {'attempted': False}
        return True, df, [], recovery_info
        
    except Exception as e:
        issues.append(f"Failed to read worksheet: {str(e)}")
        
        # Final attempt at recovery
        recovery_success, recovered_df, recovery_notes = attempt_worksheet_recovery(file_path, sheet_name)
        
        if recovery_success:
            recovery_info = {
                'attempted': True,
                'successful': True,
                'notes': recovery_notes,
                'original_issues': issues.copy()
            }
            return True, recovered_df, [], recovery_info
        else:
            recovery_info = {
                'attempted': True,
                'successful': False,
                'notes': recovery_notes
            }
            return False, None, issues, recovery_info

def calculate_worksheet_data_score(df):
    """
    Calculate a data score for worksheet to determine which has the most data
    Returns: (data_score, total_cells, non_null_cells, rows, columns)
    """
    rows, columns = df.shape
    total_cells = rows * columns
    non_null_cells = df.count().sum()  # Total non-null values
    
    # Primary score: non-null cells (actual data)
    # Secondary score: total rows (more data points)
    # Tertiary score: columns (more dimensions)
    data_score = (non_null_cells * 1000) + (rows * 10) + columns
    
    return data_score, total_cells, non_null_cells, rows, columns

def filter_and_read_clean_worksheets(file_path):
    """
    Enhanced version: Read Excel file and return only properly formatted worksheets with recovery attempts
    """
    try:
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names
        print(f"  Found {len(sheet_names)} worksheets: {sheet_names}")
        
        clean_worksheets = {}
        skipped_worksheets = {}
        worksheet_scores = {}
        recovery_log = {}
        
        for sheet_name in sheet_names:
            print(f"\n  Evaluating worksheet: '{sheet_name}'")
            
            is_valid, df, issues, recovery_info = is_worksheet_properly_formatted(file_path, sheet_name)
            recovery_log[sheet_name] = recovery_info
            
            if is_valid and df is not None:
                # Calculate data score for this worksheet
                score, total_cells, non_null_cells, rows, cols = calculate_worksheet_data_score(df)
                worksheet_scores[sheet_name] = {
                    'score': score,
                    'total_cells': total_cells,
                    'non_null_cells': non_null_cells,
                    'rows': rows,
                    'columns': cols,
                    'data_density': round((non_null_cells / total_cells) * 100, 1) if total_cells > 0 else 0
                }
                
                clean_worksheets[sheet_name] = df
                
                # Status message
                status = "INCLUDED"
                if recovery_info.get('attempted') and recovery_info.get('successful'):
                    status += " (RECOVERED)"
                
                print(f"    {status}: {rows} rows × {cols} columns")
                print(f"    Data score: {score:,} ({non_null_cells:,} non-null cells, {worksheet_scores[sheet_name]['data_density']}% density)")
                
                if recovery_info.get('attempted'):
                    print(f"    Recovery notes: {'; '.join(recovery_info.get('notes', []))}")
                
                print(f"    Columns: {list(df.columns[:5])}{'...' if len(df.columns) > 5 else ''}")
                
            else:
                skipped_worksheets[sheet_name] = issues
                print(f"    SKIPPED due to issues:")
                for issue in issues:
                    print(f"      • {issue}")
                
                if recovery_info.get('attempted') and not recovery_info.get('successful'):
                    print(f"    Recovery attempted but failed: {'; '.join(recovery_info.get('notes', []))}")
        
        # Summary
        recovered_count = sum(1 for info in recovery_log.values() 
                             if info.get('attempted') and info.get('successful'))
        
        print(f"\n  PROCESSING SUMMARY:")
        print(f"    Clean worksheets: {len(clean_worksheets)} - {list(clean_worksheets.keys())}")
        print(f"    Skipped worksheets: {len(skipped_worksheets)} - {list(skipped_worksheets.keys())}")
        if recovered_count > 0:
            print(f"    Recovered worksheets: {recovered_count}")
        
        return clean_worksheets, skipped_worksheets, worksheet_scores
        
    except Exception as e:
        print(f"  Error processing Excel file: {e}")
        return {}, {}, {}

def create_worksheet_report(clean_worksheets, skipped_worksheets, worksheet_scores, active_worksheet):
    """
    Enhanced version: Create a user-friendly report about worksheet processing including recovery info
    """
    report = []
    
    if clean_worksheets:
        report.append(f"  Successfully loaded {len(clean_worksheets)} worksheet(s):")
        
        if len(clean_worksheets) > 1:
            # Sort by data score for display
            sorted_sheets = sorted(worksheet_scores.items(), 
                                 key=lambda x: x[1]['score'], reverse=True)
            
            for i, (name, scores) in enumerate(sorted_sheets):
                active_marker = " (ACTIVE - most data)" if name == active_worksheet else ""
                report.append(f"    • '{name}': {scores['rows']:,} rows × {scores['columns']} columns, "
                             f"{scores['non_null_cells']:,} data cells{active_marker}")
        else:
            # Single worksheet
            name = list(clean_worksheets.keys())[0]
            df = clean_worksheets[name]
            report.append(f"    • '{name}': {df.shape[0]:,} rows × {df.shape[1]} columns (ACTIVE)")
    
    if skipped_worksheets:
        report.append(f"\n  Skipped {len(skipped_worksheets)} worksheet(s) due to formatting issues:")
        for name, issues in skipped_worksheets.items():
            main_issue = issues[0] if issues else "Unknown issue"
            report.append(f"    • '{name}': {main_issue}")
    
    return "\n".join(report)


def create_worksheet_report_with_exclusions(final_worksheets, skipped_worksheets, excluded_worksheets, worksheet_scores,
                                            active_worksheet):
    """
    Enhanced version: Create a comprehensive report about worksheet processing including exclusions
    """
    report = []

    if final_worksheets:
        report.append(f"  Successfully processed {len(final_worksheets)} worksheet(s):")

        if len(final_worksheets) > 1:
            # Sort by data score for display
            available_scores = {name: scores for name, scores in worksheet_scores.items() if name in final_worksheets}
            sorted_sheets = sorted(available_scores.items(),
                                   key=lambda x: x[1]['score'], reverse=True)

            for i, (name, scores) in enumerate(sorted_sheets):
                active_marker = " (ACTIVE - most data)" if name == active_worksheet else ""
                report.append(f"    • '{name}': {scores['rows']:,} rows × {scores['columns']} columns, "
                              f"{scores['non_null_cells']:,} data cells{active_marker}")
        else:
            # Single worksheet
            name = list(final_worksheets.keys())[0]
            df = final_worksheets[name]
            report.append(f"    • '{name}': {df.shape[0]:,} rows × {df.shape[1]} columns (ACTIVE)")

    # Report exclusions with categories
    total_excluded = len(skipped_worksheets) + len(excluded_worksheets)
    if total_excluded > 0:
        report.append(f"\n  Excluded {total_excluded} worksheet(s):")

        if skipped_worksheets:
            report.append(f"    {len(skipped_worksheets)} skipped due to formatting issues:")
            for name, issues in list(skipped_worksheets.items())[:3]:  # Show first 3
                main_issue = issues[0] if issues else "Unknown formatting issue"
                report.append(f"      • '{name}': {main_issue}")
            if len(skipped_worksheets) > 3:
                report.append(f"      • ... and {len(skipped_worksheets) - 3} more with formatting issues")

        if excluded_worksheets:
            report.append(f"    {len(excluded_worksheets)} excluded due to analysis issues:")
            for name, error in list(excluded_worksheets.items())[:3]:  # Show first 3
                report.append(f"      • '{name}': {error}")
            if len(excluded_worksheets) > 3:
                report.append(f"      • ... and {len(excluded_worksheets) - 3} more with analysis issues")

    # Add guidance if many worksheets were excluded
    if total_excluded > len(final_worksheets):
        report.append(f"\n  Many worksheets were excluded. Consider:")
        report.append(f"    . Simplifying worksheet formatting")
        report.append(f"    . Ensuring data is in simple table format")
        report.append(f"    . Removing charts, pivot tables, and complex structures")
            
    return "\n".join(report)


def validate_worksheet_for_analysis(df, sheet_name):
    """
    Validate if a worksheet DataFrame is suitable for AI analysis.

    Args:
        df: pandas DataFrame to validate
        sheet_name: Name of the worksheet for error reporting

    Returns:
        (is_valid, error_message): Tuple indicating if valid and any error message
    """
    try:
        # Check 1: DataFrame must not be empty
        if df.empty:
            return False, "Worksheet is empty"

        # Check 2: Must have reasonable dimensions
        rows, cols = df.shape
        if rows < 1:
            return False, "No data rows found"
        if cols < 1:
            return False, "No columns found"

        # Check 3: Must have valid column names
        if not df.columns.tolist():
            return False, "No column headers found"

        # Check 4: Columns should not be all unnamed
        unnamed_cols = [col for col in df.columns if str(col).startswith('Unnamed')]
        if len(unnamed_cols) == len(df.columns):
            return False, "All columns are unnamed - likely a formatting issue"

        # Check 5: Must have some actual data (not all NaN)
        non_null_count = df.count().sum()
        if non_null_count == 0:
            return False, "All cells are empty"

        # Check 6: Data density should be reasonable
        total_cells = rows * cols
        data_density = non_null_count / total_cells
        if data_density < 0.01:  # Less than 1% data
            return False, f"Data too sparse ({data_density:.1%} filled)"

        # Check 7: Should have at least some recognizable data types
        dtypes = df.dtypes
        if all(str(dtype) == 'object' and df[col].dropna().empty for col in df.columns):
            return False, "No recognizable data types found"

        # If we reach here, the worksheet should be analyzable
        return True, ""

    except Exception as e:
        return False, f"Validation error: {str(e)}"


def get_worksheet_analysis_capability_score(df):
    """
    Calculate a score indicating how suitable a worksheet is for analysis.
    Higher scores indicate better suitability.

    Args:
        df: pandas DataFrame to score

    Returns:
        float: Score between 0 and 1 (1 = perfect for analysis)
    """
    try:
        if df.empty:
            return 0.0

        score = 0.0

        # Factor 1: Data density (30% of score)
        rows, cols = df.shape
        total_cells = rows * cols
        non_null_count = df.count().sum()
        data_density = non_null_count / total_cells if total_cells > 0 else 0
        score += min(data_density * 2, 1.0) * 0.3  # Cap at 1.0, weight 30%

        # Factor 2: Column name quality (20% of score)
        total_cols = len(df.columns)
        unnamed_cols = len([col for col in df.columns if str(col).startswith('Unnamed')])
        named_ratio = (total_cols - unnamed_cols) / total_cols if total_cols > 0 else 0
        score += named_ratio * 0.2

        # Factor 3: Data type diversity (20% of score)
        numeric_cols = len(df.select_dtypes(include=[np.number]).columns)
        text_cols = len(df.select_dtypes(include=['object', 'string']).columns)
        datetime_cols = len(df.select_dtypes(include=['datetime']).columns)

        type_diversity = min((numeric_cols + text_cols + datetime_cols) / total_cols, 1.0) if total_cols > 0 else 0
        score += type_diversity * 0.2

        # Factor 4: Data volume (15% of score)
        volume_score = min(rows / 100, 1.0)  # Good if at least 100 rows
        score += volume_score * 0.15

        # Factor 5: Column count (15% of score)
        col_score = min(cols / 10, 1.0)  # Good if at least 10 columns
        score += col_score * 0.15

        return min(score, 1.0)  # Cap at 1.0

    except Exception:
        return 0.0